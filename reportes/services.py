from dataclasses import dataclass
from datetime import datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from html import escape
from io import BytesIO
from zoneinfo import ZoneInfo

from django.conf import settings
from django.db.models import Count, Exists, OuterRef, Q, Sum
from django.utils import timezone

from catalogo.models import Familia, Producto
from empresas.models import SucursalEmpresa
from pagos.models import Pago
from pedidos.models import DetallePedido, DetallePedidoComponente, Pedido
from pedidos.prefacturas import (
    _ajustar_fuente,
    _color,
    _color_contraste,
    _color_tenue,
    _contenido_logo_empresa,
    _imagen_logo,
    _recortar_margenes_logo,
    _texto_contacto_empresa,
)


ZERO = Decimal("0.00")
MONEY_QUANTIZER = Decimal("0.01")
PERCENT_QUANTIZER = Decimal("0.1")
MESES = (
    "",
    "Ene",
    "Feb",
    "Mar",
    "Abr",
    "May",
    "Jun",
    "Jul",
    "Ago",
    "Sep",
    "Oct",
    "Nov",
    "Dic",
)
ORDEN_ESTADOS = ("pagado", "pendiente", "rechazado", "cancelado")


def formatear_monto(valor):
    return str((valor or ZERO).quantize(MONEY_QUANTIZER, rounding=ROUND_HALF_UP))


def _limpiar_celda(valor):
    if valor is None:
        return ""
    texto = str(valor)
    if texto.startswith(("=", "+", "-", "@")):
        return f"'{texto}"
    return texto


@dataclass(frozen=True)
class ReporteTabular:
    empresa: object
    titulo: str
    metadata: tuple
    totales: tuple
    encabezados: tuple
    filas: tuple


class ReporteVentasService:
    def __init__(
        self,
        empresa,
        fecha_desde,
        fecha_hasta,
        agrupacion="dia",
        ciudad="",
        sucursal_id=None,
        examen_id=None,
        familia_id=None,
    ):
        self.empresa = empresa
        self.fecha_desde = fecha_desde
        self.fecha_hasta = fecha_hasta
        self.agrupacion = agrupacion
        self.ciudad = ciudad.strip() if ciudad else ""
        self.sucursal_id = sucursal_id
        self.examen_id = examen_id
        self.familia_id = familia_id
        self.zona_horaria = ZoneInfo(settings.TIME_ZONE)
        self.inicio, self.fin_exclusivo = self._limites(
            fecha_desde,
            fecha_hasta,
        )

    def construir_resumen(self, comparar_periodo_anterior=False):
        pedidos = self._pedidos_periodo(self.fecha_desde, self.fecha_hasta)
        confirmados = self._pedidos_confirmados(pedidos)
        totales = self._totales_confirmados(confirmados)
        estados, acumulados_estados = self._estados(pedidos)
        pagos_por_metodo = self._pagos_por_metodo(pedidos)
        pendientes_por_metodo = self._pendientes_por_metodo(pedidos)

        variacion_ingresos = None
        variacion_ventas = None
        if comparar_periodo_anterior:
            dias_periodo = (self.fecha_hasta - self.fecha_desde).days + 1
            anterior_hasta = self.fecha_desde - timedelta(days=1)
            anterior_desde = anterior_hasta - timedelta(days=dias_periodo - 1)
            anteriores = self._pedidos_confirmados(
                self._pedidos_periodo(anterior_desde, anterior_hasta)
            )
            totales_anteriores = self._totales_confirmados(anteriores)
            variacion_ingresos = self._variacion(
                totales["ingresos"],
                totales_anteriores["ingresos"],
            )
            variacion_ventas = self._variacion(
                Decimal(totales["ventas"]),
                Decimal(totales_anteriores["ventas"]),
            )

        return {
            "periodo": {
                "fecha_desde": self.fecha_desde.isoformat(),
                "fecha_hasta": self.fecha_hasta.isoformat(),
            },
            "filtros_aplicados": self._filtros_aplicados(),
            "resumen": {
                "ingresos_confirmados": formatear_monto(totales["ingresos"]),
                "ventas_confirmadas": totales["ventas"],
                "ticket_promedio": formatear_monto(totales["ticket_promedio"]),
                "subtotal": formatear_monto(totales["subtotal"]),
                "descuentos": formatear_monto(totales["descuentos"]),
                "impuestos": formatear_monto(totales["impuestos"]),
                "envios": formatear_monto(totales["envios"]),
                "monto_pendiente": formatear_monto(
                    acumulados_estados["pendiente"]["monto"]
                ),
                "pedidos_pendientes": acumulados_estados["pendiente"]["cantidad"],
                "pagos_por_metodo": pagos_por_metodo,
                "pendientes_por_metodo": pendientes_por_metodo,
                "variacion_ingresos_porcentaje": variacion_ingresos,
                "variacion_ventas_porcentaje": variacion_ventas,
            },
            "serie": self._serie(confirmados),
            "estados": estados,
            "productos_mas_vendidos": self._productos_mas_vendidos(confirmados),
        }

    def construir_tabla(self, tipo):
        resumen = self.construir_resumen(comparar_periodo_anterior=False)
        metadata = [
            ("Empresa", self.empresa.nombre),
            (
                "Periodo",
                f"{self.fecha_desde.isoformat()} a {self.fecha_hasta.isoformat()}",
            ),
        ]
        metadata.extend(self._metadata_filtros())
        metadata = tuple(metadata)
        resumen_totales = resumen["resumen"]
        totales = (
            ("Ingresos confirmados", resumen_totales["ingresos_confirmados"]),
            ("Ventas confirmadas", resumen_totales["ventas_confirmadas"]),
            ("Ticket promedio", resumen_totales["ticket_promedio"]),
            ("Subtotal", resumen_totales["subtotal"]),
            ("Descuentos", resumen_totales["descuentos"]),
            ("Impuestos", resumen_totales["impuestos"]),
            ("Envios", resumen_totales["envios"]),
            ("Monto pendiente", resumen_totales["monto_pendiente"]),
            ("Pedidos pendientes", resumen_totales["pedidos_pendientes"]),
        )

        if tipo == "resumen":
            filas = []
            for producto in resumen["productos_mas_vendidos"]:
                filas.append(
                    (
                        "Producto",
                        producto["codigo"],
                        producto["nombre"],
                        producto["cantidad"],
                        producto["ingresos"],
                    )
                )
            return ReporteTabular(
                empresa=self.empresa,
                titulo="Resumen comercial",
                metadata=metadata,
                totales=totales,
                encabezados=("Seccion", "Codigo", "Concepto", "Cantidad", "Monto"),
                filas=tuple(filas),
            )

        if tipo == "ventas":
            pedidos = self._pedidos_periodo(
                self.fecha_desde,
                self.fecha_hasta,
            ).select_related("usuario")
            filas = tuple(
                (
                    self._fecha_local(pedido.fecha_creacion),
                    pedido.numero,
                    self._estado_pedido(pedido),
                    pedido.usuario.email,
                    formatear_monto(pedido.subtotal),
                    formatear_monto(pedido.descuento_total),
                    formatear_monto(pedido.impuesto),
                    formatear_monto(pedido.envio),
                    formatear_monto(pedido.total),
                )
                for pedido in pedidos.order_by("fecha_creacion", "id")
            )
            return ReporteTabular(
                empresa=self.empresa,
                titulo="Detalle de ventas",
                metadata=metadata,
                totales=totales,
                encabezados=(
                    "Fecha",
                    "Pedido",
                    "Estado",
                    "Cliente",
                    "Subtotal",
                    "Descuentos",
                    "Impuestos",
                    "Envio",
                    "Total",
                ),
                filas=filas,
            )

        if tipo == "pagos":
            return self._tabla_pagos(metadata)

        if tipo == "impuestos":
            return self._tabla_impuestos(metadata, totales)

        if tipo == "sucursales":
            return self._tabla_sucursales(metadata)

        return self._tabla_familias(metadata)

    def _pedidos_periodo(self, fecha_desde, fecha_hasta):
        inicio, fin_exclusivo = self._limites(fecha_desde, fecha_hasta)
        aprobados = Pago.objects.filter(
            pedido_id=OuterRef("pk"),
            estado=Pago.Estado.APROBADO,
        )
        pendientes = Pago.objects.filter(
            pedido_id=OuterRef("pk"),
            estado=Pago.Estado.PENDIENTE,
        )
        rechazados = Pago.objects.filter(
            pedido_id=OuterRef("pk"),
            estado=Pago.Estado.RECHAZADO,
        )
        pedidos = Pedido.objects.filter(
            empresa=self.empresa,
            fecha_creacion__gte=inicio,
            fecha_creacion__lt=fin_exclusivo,
        ).annotate(
            tiene_pago_aprobado=Exists(aprobados),
            tiene_pago_pendiente=Exists(pendientes),
            tiene_pago_rechazado=Exists(rechazados),
        )
        if self.ciudad:
            pedidos = pedidos.filter(
                Q(municipio_entrega__iexact=self.ciudad)
                | Q(sucursal_pago__ciudad__iexact=self.ciudad)
            )
        if self.sucursal_id:
            pedidos = pedidos.filter(sucursal_pago_id=self.sucursal_id)
        if self.examen_id:
            pedidos = pedidos.filter(
                Q(detalles__producto_id=self.examen_id)
                | Q(detalles__componentes__producto_id=self.examen_id)
            )
        if self.familia_id:
            pedidos = pedidos.filter(
                Q(detalles__producto__familia_id=self.familia_id)
                | Q(
                    detalles__componentes__producto__familia_id=self.familia_id
                )
            )
        return pedidos.distinct()

    def _pedidos_confirmados(self, pedidos):
        return pedidos.filter(
            Q(estado_pago=Pedido.EstadoPago.PAGADO)
            | Q(tiene_pago_aprobado=True)
        )

    def _totales_confirmados(self, confirmados):
        agregado = confirmados.aggregate(
            ingresos=Sum("total"),
            subtotal=Sum("subtotal"),
            descuentos=Sum("descuento_total"),
            impuestos=Sum("impuesto"),
            envios=Sum("envio"),
            ventas=Count("id"),
        )
        ventas = agregado["ventas"] or 0
        ingresos = agregado["ingresos"] or ZERO
        return {
            "ingresos": ingresos,
            "subtotal": agregado["subtotal"] or ZERO,
            "descuentos": agregado["descuentos"] or ZERO,
            "impuestos": agregado["impuestos"] or ZERO,
            "envios": agregado["envios"] or ZERO,
            "ventas": ventas,
            "ticket_promedio": ingresos / ventas if ventas else ZERO,
        }

    def _pagos_por_metodo(self, pedidos):
        acumulados = {
            Pago.Metodo.SUCURSAL: {"cantidad": 0, "monto": ZERO},
            Pago.Metodo.EN_LINEA: {"cantidad": 0, "monto": ZERO},
        }
        pagos = (
            Pago.objects.filter(
                empresa=self.empresa,
                pedido__empresa=self.empresa,
                pedido__in=pedidos,
                estado=Pago.Estado.APROBADO,
                metodo__in=acumulados,
            )
            .order_by("pedido_id", "id")
            .values_list("pedido_id", "metodo", "monto")
        )
        pedidos_contados = set()
        for pedido_id, metodo, monto in pagos:
            if pedido_id in pedidos_contados:
                continue
            pedidos_contados.add(pedido_id)
            acumulados[metodo]["cantidad"] += 1
            acumulados[metodo]["monto"] += monto

        return {
            metodo: {
                "cantidad": valores["cantidad"],
                "monto": formatear_monto(valores["monto"]),
            }
            for metodo, valores in acumulados.items()
        }

    def _pendientes_por_metodo(self, pedidos):
        acumulados = {
            Pedido.MetodoPago.SUCURSAL: {"cantidad": 0, "monto": ZERO},
            Pedido.MetodoPago.EN_LINEA: {"cantidad": 0, "monto": ZERO},
            "sin_metodo": {"cantidad": 0, "monto": ZERO},
        }
        for pedido in pedidos.only("estado_pago", "metodo_pago", "total"):
            if (
                pedido.estado_pago != Pedido.EstadoPago.PENDIENTE
                or pedido.tiene_pago_aprobado
            ):
                continue
            metodo = (
                pedido.metodo_pago
                if pedido.metodo_pago in {
                    Pedido.MetodoPago.SUCURSAL,
                    Pedido.MetodoPago.EN_LINEA,
                }
                else "sin_metodo"
            )
            acumulados[metodo]["cantidad"] += 1
            acumulados[metodo]["monto"] += pedido.total

        return {
            metodo: {
                "cantidad": valores["cantidad"],
                "monto": formatear_monto(valores["monto"]),
            }
            for metodo, valores in acumulados.items()
        }

    def _estados(self, pedidos):
        acumulados = {
            estado: {"cantidad": 0, "monto": ZERO}
            for estado in ORDEN_ESTADOS
        }
        for pedido in pedidos.only("estado_pago", "total"):
            estado = self._estado_pedido(pedido)
            acumulados[estado]["cantidad"] += 1
            acumulados[estado]["monto"] += pedido.total

        estados = [
            {
                "estado": estado,
                "cantidad": acumulados[estado]["cantidad"],
                "monto": formatear_monto(acumulados[estado]["monto"]),
            }
            for estado in ORDEN_ESTADOS
            if acumulados[estado]["cantidad"]
        ]
        return estados, acumulados

    def _estado_pedido(self, pedido):
        if (
            pedido.estado_pago == Pedido.EstadoPago.PAGADO
            or pedido.tiene_pago_aprobado
        ):
            return "pagado"
        if pedido.estado_pago == Pedido.EstadoPago.CANCELADO:
            return "cancelado"
        if pedido.estado_pago == Pedido.EstadoPago.RECHAZADO:
            return "rechazado"
        if pedido.tiene_pago_pendiente:
            return "pendiente"
        if pedido.tiene_pago_rechazado:
            return "rechazado"
        return "pendiente"

    def _serie(self, confirmados):
        acumulados = {}
        for fecha_creacion, total in confirmados.values_list(
            "fecha_creacion",
            "total",
        ):
            fecha_local = timezone.localtime(
                fecha_creacion,
                self.zona_horaria,
            ).date()
            clave = (
                fecha_local
                if self.agrupacion == "dia"
                else fecha_local.replace(day=1)
            )
            if clave not in acumulados:
                acumulados[clave] = {"ingresos": ZERO, "ventas": 0}
            acumulados[clave]["ingresos"] += total
            acumulados[clave]["ventas"] += 1

        return [
            {
                "periodo": self._periodo_serie(fecha),
                "etiqueta": self._etiqueta_serie(fecha),
                "ingresos": formatear_monto(
                    acumulados.get(fecha, {}).get("ingresos", ZERO)
                ),
                "ventas": acumulados.get(fecha, {}).get("ventas", 0),
            }
            for fecha in self._periodos_serie()
        ]

    def _periodos_serie(self):
        if self.agrupacion == "dia":
            actual = self.fecha_desde
            while actual <= self.fecha_hasta:
                yield actual
                actual += timedelta(days=1)
            return

        actual = self.fecha_desde.replace(day=1)
        ultimo = self.fecha_hasta.replace(day=1)
        while actual <= ultimo:
            yield actual
            actual = (
                actual.replace(year=actual.year + 1, month=1)
                if actual.month == 12
                else actual.replace(month=actual.month + 1)
            )

    def _periodo_serie(self, fecha):
        return fecha.isoformat() if self.agrupacion == "dia" else fecha.strftime("%Y-%m")

    def _etiqueta_serie(self, fecha):
        varios_anios = self.fecha_desde.year != self.fecha_hasta.year
        if self.agrupacion == "mes":
            return f"{MESES[fecha.month]} {fecha.year}" if varios_anios else MESES[fecha.month]
        etiqueta = f"{fecha.day:02d} {MESES[fecha.month]}"
        return f"{etiqueta} {fecha.year}" if varios_anios else etiqueta

    def _productos_mas_vendidos(self, confirmados):
        productos = (
            self._detalles_filtrados(confirmados)
            .values("codigo_articulo", "nombre_articulo")
            .annotate(
                cantidad=Sum("cantidad"),
                ingresos=Sum("subtotal_final"),
            )
            .order_by("-cantidad", "-ingresos", "nombre_articulo")[:10]
        )
        return [
            {
                "codigo": producto["codigo_articulo"],
                "nombre": producto["nombre_articulo"],
                "cantidad": producto["cantidad"],
                "ingresos": formatear_monto(producto["ingresos"]),
            }
            for producto in productos
        ]

    def _detalles_filtrados(self, pedidos):
        detalles = DetallePedido.objects.filter(pedido__in=pedidos)
        if self.examen_id:
            componentes_examen = DetallePedidoComponente.objects.filter(
                detalle_id=OuterRef("pk"),
                producto_id=self.examen_id,
            )
            detalles = detalles.annotate(
                contiene_examen=Exists(componentes_examen)
            ).filter(
                Q(producto_id=self.examen_id) | Q(contiene_examen=True)
            )
        if self.familia_id:
            componentes_familia = DetallePedidoComponente.objects.filter(
                detalle_id=OuterRef("pk"),
                producto__familia_id=self.familia_id,
            )
            detalles = detalles.annotate(
                contiene_familia=Exists(componentes_familia)
            ).filter(
                Q(producto__familia_id=self.familia_id)
                | Q(contiene_familia=True)
            )
        return detalles

    def _filtros_aplicados(self):
        return {
            "ciudad": self.ciudad or None,
            "sucursal_id": self.sucursal_id,
            "examen_id": self.examen_id,
            "familia_id": self.familia_id,
        }

    def _metadata_filtros(self):
        metadata = []
        if self.ciudad:
            metadata.append(("Ciudad", self.ciudad))
        if self.sucursal_id:
            sucursal = SucursalEmpresa.objects.filter(
                pk=self.sucursal_id,
                empresa=self.empresa,
            ).first()
            metadata.append(
                ("Sucursal", sucursal.nombre if sucursal else self.sucursal_id)
            )
        if self.examen_id:
            examen = Producto.objects.filter(
                pk=self.examen_id,
                empresa=self.empresa,
            ).first()
            valor = (
                f"{examen.codigo_venta} - {examen.nombre}"
                if examen
                else self.examen_id
            )
            metadata.append(("Examen", valor))
        if self.familia_id:
            familia = Familia.objects.filter(
                pk=self.familia_id,
                empresa=self.empresa,
            ).first()
            metadata.append(
                ("Familia", familia.nombre if familia else self.familia_id)
            )
        return metadata

    def _tabla_pagos(self, metadata):
        pagos = Pago.objects.filter(
            empresa=self.empresa,
            pedido__in=self._pedidos_periodo(
                self.fecha_desde,
                self.fecha_hasta,
            ),
        ).select_related("pedido", "usuario")
        agregado = pagos.values("estado").annotate(
            cantidad=Count("id"),
            monto=Sum("monto"),
        )
        por_estado = {
            item["estado"]: item
            for item in agregado
        }
        totales = tuple(
            (
                f"Pagos {estado}",
                f"{por_estado.get(estado, {}).get('cantidad', 0)} / "
                f"{formatear_monto(por_estado.get(estado, {}).get('monto', ZERO))}",
            )
            for estado in (
                Pago.Estado.APROBADO,
                Pago.Estado.PENDIENTE,
                Pago.Estado.RECHAZADO,
                Pago.Estado.CANCELADO,
            )
        )
        filas = tuple(
            (
                self._fecha_local(pago.fecha_creacion),
                pago.pedido.numero,
                str(pago.referencia),
                pago.proveedor,
                pago.estado,
                pago.usuario.email,
                formatear_monto(pago.monto),
                self._fecha_local(pago.fecha_confirmacion)
                if pago.fecha_confirmacion
                else "",
            )
            for pago in pagos.order_by("fecha_creacion", "id")
        )
        return ReporteTabular(
            empresa=self.empresa,
            titulo="Detalle de pagos",
            metadata=metadata,
            totales=totales,
            encabezados=(
                "Fecha",
                "Pedido",
                "Referencia",
                "Proveedor",
                "Estado",
                "Cliente",
                "Monto",
                "Fecha confirmacion",
            ),
            filas=filas,
        )

    def _tabla_impuestos(self, metadata, totales_resumen):
        confirmados = self._pedidos_confirmados(
            self._pedidos_periodo(self.fecha_desde, self.fecha_hasta)
        ).select_related("usuario")
        filas = tuple(
            (
                self._fecha_local(pedido.fecha_creacion),
                pedido.numero,
                pedido.usuario.email,
                formatear_monto(pedido.subtotal - pedido.descuento_total),
                str(pedido.tasa_impuesto),
                formatear_monto(pedido.impuesto),
                formatear_monto(pedido.total),
            )
            for pedido in confirmados.order_by("fecha_creacion", "id")
        )
        return ReporteTabular(
            empresa=self.empresa,
            titulo="Detalle de impuestos",
            metadata=metadata,
            totales=totales_resumen,
            encabezados=(
                "Fecha",
                "Pedido",
                "Cliente",
                "Base imponible",
                "Tasa",
                "Impuesto",
                "Total",
            ),
            filas=filas,
        )

    def _tabla_sucursales(self, metadata):
        pedidos = (
            self._pedidos_periodo(self.fecha_desde, self.fecha_hasta)
            .filter(
                metodo_pago=Pedido.MetodoPago.SUCURSAL,
                sucursal_pago__isnull=False,
            )
            .select_related("sucursal_pago")
            .order_by("fecha_creacion", "id")
        )
        acumulados = {}
        personas_globales = set()
        monto_total = ZERO
        for pedido in pedidos:
            sucursal = pedido.sucursal_pago
            datos = acumulados.setdefault(
                sucursal.pk,
                {
                    "ciudad": sucursal.ciudad or "Sin ciudad registrada",
                    "sucursal": sucursal.nombre,
                    "personas": set(),
                    "selecciones": 0,
                    "pagados": 0,
                    "pendientes": 0,
                    "otros": 0,
                    "monto": ZERO,
                },
            )
            datos["personas"].add(pedido.usuario_id)
            personas_globales.add(pedido.usuario_id)
            datos["selecciones"] += 1
            estado = self._estado_pedido(pedido)
            if estado == "pagado":
                datos["pagados"] += 1
            elif estado == "pendiente":
                datos["pendientes"] += 1
            else:
                datos["otros"] += 1
            datos["monto"] += pedido.total
            monto_total += pedido.total

        ordenados = sorted(
            acumulados.values(),
            key=lambda item: (
                -len(item["personas"]),
                -item["selecciones"],
                item["sucursal"].lower(),
            ),
        )
        filas = tuple(
            (
                item["ciudad"],
                item["sucursal"],
                len(item["personas"]),
                item["selecciones"],
                item["pagados"],
                item["pendientes"],
                item["otros"],
                formatear_monto(item["monto"]),
            )
            for item in ordenados
        )
        totales = (
            ("Sucursales seleccionadas", len(acumulados)),
            ("Personas unicas", len(personas_globales)),
            ("Selecciones de sucursal", sum(i["selecciones"] for i in ordenados)),
            ("Pedidos pagados", sum(i["pagados"] for i in ordenados)),
            ("Pedidos pendientes", sum(i["pendientes"] for i in ordenados)),
            ("Monto seleccionado", formatear_monto(monto_total)),
        )
        return ReporteTabular(
            empresa=self.empresa,
            titulo="Visitas estimadas por sucursal",
            metadata=metadata,
            totales=totales,
            encabezados=(
                "Ciudad",
                "Sucursal",
                "Personas",
                "Selecciones",
                "Pagados",
                "Pendientes",
                "Otros",
                "Monto",
            ),
            filas=filas,
        )

    def _tabla_familias(self, metadata):
        confirmados = self._pedidos_confirmados(
            self._pedidos_periodo(self.fecha_desde, self.fecha_hasta)
        )
        detalles = self._detalles_filtrados(confirmados).filter(
            producto__isnull=False
        )
        ventas = {
            item["producto_id"]: item
            for item in detalles.values("producto_id").annotate(
                pedidos=Count("pedido_id", distinct=True),
                cantidad=Sum("cantidad"),
                ingresos=Sum("subtotal_final"),
            )
        }
        productos = Producto.objects.filter(empresa=self.empresa).select_related(
            "familia"
        )
        if self.familia_id:
            productos = productos.filter(familia_id=self.familia_id)
        if self.examen_id:
            productos = productos.filter(pk=self.examen_id)
        productos = productos.order_by(
            "familia__orden",
            "familia__nombre",
            "nombre",
        )

        filas = tuple(
            (
                producto.familia.nombre,
                producto.codigo_venta,
                producto.nombre,
                "Activo" if producto.activo else "Inactivo",
                ventas.get(producto.pk, {}).get("pedidos", 0),
                ventas.get(producto.pk, {}).get("cantidad", 0),
                formatear_monto(
                    ventas.get(producto.pk, {}).get("ingresos", ZERO)
                ),
            )
            for producto in productos
        )
        pedidos_globales = detalles.values("pedido_id").distinct().count()
        cantidad_total = sum(
            item.get("cantidad") or 0
            for item in ventas.values()
        )
        ingresos_totales = sum(
            (item.get("ingresos") or ZERO for item in ventas.values()),
            ZERO,
        )
        familias_incluidas = len({fila[0] for fila in filas})
        totales = (
            ("Familias incluidas", familias_incluidas),
            ("Productos incluidos", len(filas)),
            ("Productos con ventas", len(ventas)),
            ("Pedidos confirmados", pedidos_globales),
            ("Unidades vendidas", cantidad_total),
            ("Ingresos directos", formatear_monto(ingresos_totales)),
        )
        return ReporteTabular(
            empresa=self.empresa,
            titulo="Detalle de productos por familia",
            metadata=metadata,
            totales=totales,
            encabezados=(
                "Familia",
                "Codigo",
                "Producto o examen",
                "Estado",
                "Pedidos",
                "Cantidad",
                "Ingresos",
            ),
            filas=filas,
        )

    def _limites(self, fecha_desde, fecha_hasta):
        inicio = datetime.combine(fecha_desde, time.min, self.zona_horaria)
        fin_exclusivo = datetime.combine(
            fecha_hasta + timedelta(days=1),
            time.min,
            self.zona_horaria,
        )
        return inicio, fin_exclusivo

    def _fecha_local(self, valor):
        return timezone.localtime(valor, self.zona_horaria).strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    def _variacion(self, actual, anterior):
        if not anterior:
            return None
        porcentaje = ((actual - anterior) / anterior) * Decimal("100")
        return float(porcentaje.quantize(PERCENT_QUANTIZER, rounding=ROUND_HALF_UP))


def exportar_xlsx(tabla):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    empresa = tabla.empresa

    def color_excel(valor, predeterminado):
        texto = str(valor or predeterminado).strip().lstrip("#").upper()
        return texto if len(texto) == 6 else predeterminado.lstrip("#").upper()

    def contraste(valor):
        rojo = int(valor[0:2], 16)
        verde = int(valor[2:4], 16)
        azul = int(valor[4:6], 16)
        luminancia = (0.299 * rojo) + (0.587 * verde) + (0.114 * azul)
        return "000000" if luminancia > 158 else "FFFFFF"

    def aclarar(valor, proporcion=0.9):
        componentes = [int(valor[indice : indice + 2], 16) for indice in (0, 2, 4)]
        claros = [
            round(componente + ((255 - componente) * proporcion))
            for componente in componentes
        ]
        return "".join(f"{componente:02X}" for componente in claros)

    principal = color_excel(empresa.color_principal, "D1393D")
    secundario = color_excel(empresa.color_secundario, "E94A51")
    acento = color_excel(empresa.color_acento, "2D4B77")
    texto = color_excel(empresa.color_texto, "111827")
    fondo_suave = aclarar(acento, 0.92)
    borde_color = aclarar(acento, 0.7)
    borde_fino = Side(style="thin", color=borde_color)
    borde = Border(
        left=borde_fino,
        right=borde_fino,
        top=borde_fino,
        bottom=borde_fino,
    )

    encabezados_numericos = {
        "cantidad",
        "personas",
        "selecciones",
        "pagados",
        "pendientes",
        "otros",
        "pedidos",
        "ingresos",
        "subtotal",
        "descuentos",
        "impuestos",
        "impuesto",
        "envio",
        "total",
        "monto",
        "tasa",
        "base imponible",
    }
    encabezados_enteros = {
        "cantidad",
        "personas",
        "selecciones",
        "pagados",
        "pendientes",
        "otros",
        "pedidos",
    }

    def convertir_numero(valor, encabezado):
        clave = str(encabezado).strip().lower()
        if clave not in encabezados_numericos:
            return _limpiar_celda(valor), None
        try:
            numero = Decimal(str(valor))
        except (ArithmeticError, ValueError):
            return _limpiar_celda(valor), None
        if clave in encabezados_enteros:
            return int(numero), "#,##0"
        if clave == "tasa":
            return float(numero), "0.00%"
        return float(numero), '#,##0.00;[Red]-#,##0.00;"-"'

    def convertir_total(etiqueta, valor):
        clave = str(etiqueta).strip().lower()
        if clave in {
            "ventas confirmadas",
            "pedidos pendientes",
            "sucursales seleccionadas",
            "personas unicas",
            "selecciones de sucursal",
            "pedidos pagados",
            "familias con ventas",
            "familias incluidas",
            "productos incluidos",
            "productos con ventas",
            "pedidos confirmados",
            "unidades vendidas",
        }:
            try:
                return int(Decimal(str(valor))), "#,##0"
            except (ArithmeticError, ValueError):
                return _limpiar_celda(valor), None
        try:
            return float(Decimal(str(valor))), '#,##0.00;[Red]-#,##0.00;"-"'
        except (ArithmeticError, ValueError):
            return _limpiar_celda(valor), None

    libro = Workbook()
    libro.properties.creator = empresa.nombre
    libro.properties.title = tabla.titulo
    libro.properties.subject = "Reporte comercial"
    hoja = libro.active
    hoja.title = "Reporte"
    hoja.sheet_view.showGridLines = False
    hoja.sheet_view.zoomScale = 90

    cantidad_columnas = max(len(tabla.encabezados), 6)
    ultima_columna = get_column_letter(cantidad_columnas)
    inicio_reporte = max(4, cantidad_columnas - 1)
    fin_empresa = inicio_reporte - 1
    columna_inicio_reporte = get_column_letter(inicio_reporte)
    columna_fin_empresa = get_column_letter(fin_empresa)

    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=fin_empresa)
    hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=fin_empresa)
    hoja.merge_cells(
        start_row=1,
        start_column=inicio_reporte,
        end_row=1,
        end_column=cantidad_columnas,
    )
    hoja.merge_cells(
        start_row=2,
        start_column=inicio_reporte,
        end_row=2,
        end_column=cantidad_columnas,
    )
    hoja.merge_cells(f"A3:{ultima_columna}3")

    hoja["A1"] = empresa.nombre
    hoja["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=acento)
    hoja["A1"].alignment = Alignment(horizontal="center", vertical="center")
    contacto = _texto_contacto_empresa(empresa)
    hoja["A2"] = contacto
    hoja["A2"].font = Font(name="Aptos", size=8, color="4B5563")
    hoja["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        shrink_to_fit=True,
    )
    hoja[f"{columna_inicio_reporte}1"] = "REPORTE COMERCIAL"
    hoja[f"{columna_inicio_reporte}1"].font = Font(
        name="Aptos Display",
        size=15,
        bold=True,
        color=principal,
    )
    hoja[f"{columna_inicio_reporte}1"].alignment = Alignment(
        horizontal="right",
        vertical="center",
    )
    hoja[f"{columna_inicio_reporte}2"] = tabla.titulo
    hoja[f"{columna_inicio_reporte}2"].font = Font(
        name="Aptos",
        size=10,
        bold=True,
        color=texto,
    )
    hoja[f"{columna_inicio_reporte}2"].alignment = Alignment(
        horizontal="right",
        vertical="center",
    )
    hoja[f"{columna_inicio_reporte}1"].border = Border(
        left=Side(style="medium", color=secundario)
    )
    hoja[f"{columna_inicio_reporte}2"].border = Border(
        left=Side(style="medium", color=secundario)
    )

    hoja["A3"] = tabla.titulo.upper()
    hoja["A3"].font = Font(
        name="Aptos",
        size=11,
        bold=True,
        color=contraste(principal),
    )
    hoja["A3"].fill = PatternFill("solid", fgColor=principal)
    hoja["A3"].alignment = Alignment(horizontal="center", vertical="center")
    hoja.row_dimensions[1].height = 30
    hoja.row_dimensions[2].height = 20
    hoja.row_dimensions[3].height = 22

    if contenido_logo := _contenido_logo_empresa(empresa):
        try:
            contenido_logo = _recortar_margenes_logo(contenido_logo)
            imagen = ExcelImage(BytesIO(contenido_logo))
            escala = min(92 / imagen.width, 42 / imagen.height)
            imagen.width *= escala
            imagen.height *= escala
            hoja.add_image(imagen, "A1")
        except Exception:
            pass

    def combinar_bloque(fila, bloque, total_bloques=3):
        inicio = (bloque * cantidad_columnas // total_bloques) + 1
        fin = ((bloque + 1) * cantidad_columnas // total_bloques)
        if fin > inicio:
            hoja.merge_cells(
                start_row=fila,
                start_column=inicio,
                end_row=fila,
                end_column=fin,
            )
        return hoja.cell(row=fila, column=inicio)

    fila_seccion_datos = 5
    hoja.merge_cells(f"A{fila_seccion_datos}:{ultima_columna}{fila_seccion_datos}")
    hoja[f"A{fila_seccion_datos}"] = "DATOS DEL REPORTE"
    metadata = list(tabla.metadata)
    metadata.append(
        (
            "Generado",
            timezone.localtime(timezone.now(), ZoneInfo(settings.TIME_ZONE)).strftime(
                "%d/%m/%Y %H:%M"
            ),
        )
    )
    fila_metadata = fila_seccion_datos + 1
    for indice in range(0, len(metadata), 3):
        grupo = metadata[indice : indice + 3]
        for bloque, (clave, valor) in enumerate(grupo):
            etiqueta = combinar_bloque(fila_metadata, bloque)
            dato = combinar_bloque(fila_metadata + 1, bloque)
            etiqueta.value = str(clave).upper()
            dato.value = _limpiar_celda(valor)
            for celda in (etiqueta, dato):
                celda.border = borde
                celda.alignment = Alignment(vertical="center", wrap_text=True)
            etiqueta.font = Font(name="Aptos", size=8, color="6B7280")
            etiqueta.fill = PatternFill("solid", fgColor=fondo_suave)
            dato.font = Font(name="Aptos", size=10, bold=True, color=texto)
        hoja.row_dimensions[fila_metadata].height = 18
        hoja.row_dimensions[fila_metadata + 1].height = 24
        fila_metadata += 2

    fila_seccion_totales = fila_metadata + 1
    hoja.merge_cells(
        f"A{fila_seccion_totales}:{ultima_columna}{fila_seccion_totales}"
    )
    hoja[f"A{fila_seccion_totales}"] = "RESUMEN DE TOTALES"
    fila_total = fila_seccion_totales + 1
    for indice in range(0, len(tabla.totales), 3):
        grupo = tabla.totales[indice : indice + 3]
        for bloque, (clave, valor) in enumerate(grupo):
            etiqueta = combinar_bloque(fila_total, bloque)
            dato = combinar_bloque(fila_total + 1, bloque)
            valor_excel, formato = convertir_total(clave, valor)
            etiqueta.value = str(clave).upper()
            dato.value = valor_excel
            if formato:
                dato.number_format = formato
            for celda in (etiqueta, dato):
                celda.border = borde
                celda.alignment = Alignment(vertical="center", wrap_text=True)
            etiqueta.font = Font(name="Aptos", size=8, color="6B7280")
            etiqueta.fill = PatternFill("solid", fgColor=fondo_suave)
            dato.font = Font(name="Aptos", size=11, bold=True, color=texto)
        hoja.row_dimensions[fila_total].height = 18
        hoja.row_dimensions[fila_total + 1].height = 24
        fila_total += 2

    fila_seccion_detalle = fila_total + 1
    hoja.merge_cells(
        f"A{fila_seccion_detalle}:{ultima_columna}{fila_seccion_detalle}"
    )
    hoja[f"A{fila_seccion_detalle}"] = "DETALLE"
    fila_encabezados = fila_seccion_detalle + 1
    for columna, encabezado in enumerate(tabla.encabezados, start=1):
        celda = hoja.cell(row=fila_encabezados, column=columna, value=str(encabezado))
        celda.font = Font(
            name="Aptos",
            size=9,
            bold=True,
            color=contraste(acento),
        )
        celda.fill = PatternFill("solid", fgColor=acento)
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        celda.border = borde
    for columna in range(len(tabla.encabezados) + 1, cantidad_columnas + 1):
        celda = hoja.cell(row=fila_encabezados, column=columna)
        celda.fill = PatternFill("solid", fgColor=acento)
        celda.border = borde
    hoja.row_dimensions[fila_encabezados].height = 24

    primera_fila_detalle = fila_encabezados + 1
    for indice_fila, fila in enumerate(tabla.filas, start=primera_fila_detalle):
        for indice_columna, (encabezado, valor) in enumerate(
            zip(tabla.encabezados, fila),
            start=1,
        ):
            valor_excel, formato = convertir_numero(valor, encabezado)
            celda = hoja.cell(
                row=indice_fila,
                column=indice_columna,
                value=valor_excel,
            )
            if formato:
                celda.number_format = formato
            celda.font = Font(name="Aptos", size=9, color=texto)
            celda.alignment = Alignment(
                horizontal=(
                    "right"
                    if str(encabezado).strip().lower() in encabezados_numericos
                    else "left"
                ),
                vertical="top",
                wrap_text=True,
            )
            celda.border = borde
            if (indice_fila - primera_fila_detalle) % 2:
                celda.fill = PatternFill("solid", fgColor=fondo_suave)
        hoja.row_dimensions[indice_fila].height = 20

    ultima_fila = max(hoja.max_row, fila_encabezados)
    hoja.freeze_panes = f"A{primera_fila_detalle}"
    hoja.auto_filter.ref = (
        f"A{fila_encabezados}:{get_column_letter(len(tabla.encabezados))}{ultima_fila}"
    )
    hoja.print_title_rows = f"{fila_encabezados}:{fila_encabezados}"
    hoja.print_area = f"A1:{ultima_columna}{ultima_fila}"

    for indice_columna in range(1, cantidad_columnas + 1):
        valores = [
            str(hoja.cell(row=fila, column=indice_columna).value or "")
            for fila in range(fila_encabezados, ultima_fila + 1)
        ]
        ancho = max((len(valor) for valor in valores), default=10) + 2
        hoja.column_dimensions[get_column_letter(indice_columna)].width = min(
            max(ancho, 12),
            44,
        )

    for fila_seccion in (
        fila_seccion_datos,
        fila_seccion_totales,
        fila_seccion_detalle,
    ):
        celda = hoja.cell(row=fila_seccion, column=1)
        celda.font = Font(name="Aptos", size=9, bold=True, color=acento)
        celda.fill = PatternFill("solid", fgColor=aclarar(acento, 0.95))
        celda.alignment = Alignment(vertical="center")
        hoja.row_dimensions[fila_seccion].height = 20

    hoja.page_setup.orientation = hoja.ORIENTATION_LANDSCAPE
    hoja.page_setup.paperSize = hoja.PAPERSIZE_A4
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.print_options.horizontalCentered = True
    hoja.page_margins.left = 0.3
    hoja.page_margins.right = 0.3
    hoja.page_margins.top = 0.5
    hoja.page_margins.bottom = 0.5
    hoja.oddFooter.left.text = empresa.nombre
    hoja.oddFooter.center.text = contacto
    hoja.oddFooter.right.text = "Pagina &P de &N"
    hoja.oddFooter.left.size = 8
    hoja.oddFooter.center.size = 8
    hoja.oddFooter.right.size = 8

    salida = BytesIO()
    libro.save(salida)
    return salida.getvalue()


def exportar_pdf(tabla):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        KeepTogether,
        LongTable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    empresa = tabla.empresa
    pagina = landscape(A4)
    ancho_pagina, alto_pagina = pagina
    ancho_util = ancho_pagina - (28 * mm)
    color_principal = _color(empresa.color_principal, "#D1393D")
    color_secundario = _color(empresa.color_secundario, "#E94A51")
    color_acento = _color(empresa.color_acento, "#2D4B77")
    color_texto = _color(empresa.color_texto, "#111827")
    texto_principal = _color_contraste(color_principal)
    texto_acento = _color_contraste(color_acento)
    fondo_suave = _color_tenue(color_acento, 0.92)

    contenido_logo = _contenido_logo_empresa(empresa)
    if contenido_logo:
        contenido_logo = _recortar_margenes_logo(contenido_logo)
    logo = _imagen_logo(contenido_logo)

    salida = BytesIO()
    documento = SimpleDocTemplate(
        salida,
        pagesize=pagina,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=19 * mm,
        title=tabla.titulo,
        author=empresa.nombre,
        subject="Reporte comercial",
        pageCompression=0,
        invariant=1,
    )

    estilos_base = getSampleStyleSheet()
    estilos = {
        "Empresa": ParagraphStyle(
            "EmpresaReporte",
            parent=estilos_base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=19,
            textColor=color_acento,
            spaceAfter=3,
        ),
        "Contacto": ParagraphStyle(
            "ContactoReporte",
            parent=estilos_base["BodyText"],
            fontSize=7.5,
            leading=10,
            textColor=colors.HexColor("#4B5563"),
        ),
        "Documento": ParagraphStyle(
            "DocumentoReporte",
            parent=estilos_base["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=20,
            alignment=2,
            textColor=color_principal,
        ),
        "Titulo": ParagraphStyle(
            "TituloReporte",
            parent=estilos_base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=14,
            alignment=2,
            textColor=color_texto,
        ),
        "Banda": ParagraphStyle(
            "BandaReporte",
            parent=estilos_base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=10,
            leading=13,
            alignment=1,
            textColor=texto_principal,
        ),
        "Seccion": ParagraphStyle(
            "SeccionReporte",
            parent=estilos_base["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=color_acento,
            spaceBefore=4,
            spaceAfter=5,
        ),
        "Campo": ParagraphStyle(
            "CampoReporte",
            parent=estilos_base["BodyText"],
            fontSize=8,
            leading=10.5,
            textColor=color_texto,
        ),
        "Celda": ParagraphStyle(
            "CeldaReporte",
            parent=estilos_base["BodyText"],
            fontSize=6.5,
            leading=8.2,
            textColor=color_texto,
        ),
        "CeldaDerecha": ParagraphStyle(
            "CeldaDerechaReporte",
            parent=estilos_base["BodyText"],
            fontSize=6.5,
            leading=8.2,
            alignment=2,
            textColor=color_texto,
        ),
        "Encabezado": ParagraphStyle(
            "EncabezadoReporte",
            parent=estilos_base["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=6.5,
            leading=8,
            alignment=1,
            textColor=texto_acento,
        ),
    }

    identidad_empresa = []
    if logo:
        identidad_empresa.extend([logo, Spacer(1, 2 * mm)])
    identidad_empresa.append(Paragraph(escape(empresa.nombre), estilos["Empresa"]))
    contacto_encabezado = "<br/>".join(
        escape(str(valor))
        for valor in [
            empresa.direccion,
            empresa.telefono,
            empresa.correo,
            empresa.sitio_web,
        ]
        if valor
    )
    if contacto_encabezado:
        identidad_empresa.append(
            Paragraph(contacto_encabezado, estilos["Contacto"])
        )

    cabecera = Table(
        [
            [
                identidad_empresa,
                [
                    Paragraph("REPORTE COMERCIAL", estilos["Documento"]),
                    Paragraph(escape(tabla.titulo), estilos["Titulo"]),
                ],
            ]
        ],
        colWidths=[170 * mm, ancho_util - (170 * mm)],
    )
    cabecera.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("RIGHTPADDING", (0, 0), (0, 0), 5 * mm),
                ("LEFTPADDING", (1, 0), (1, 0), 4 * mm),
                ("RIGHTPADDING", (1, 0), (1, 0), 0),
                ("LINEBEFORE", (1, 0), (1, 0), 2, color_secundario),
            ]
        )
    )

    banda = Table(
        [[Paragraph(escape(tabla.titulo.upper()), estilos["Banda"])]],
        colWidths=[ancho_util],
    )
    banda.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), color_principal),
                ("BOX", (0, 0), (-1, -1), 0.8, color_principal),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    generado = timezone.localtime(
        timezone.now(),
        ZoneInfo(settings.TIME_ZONE),
    ).strftime("%d/%m/%Y %H:%M")
    metadata = [
        (clave, valor)
        for clave, valor in tabla.metadata
        if str(clave).strip().lower() not in {"empresa slug", "moneda"}
    ]
    metadata.append(("Generado", generado))

    indices_visibles = [
        indice
        for indice, encabezado in enumerate(tabla.encabezados)
        if str(encabezado).strip().lower() != "moneda"
    ]
    encabezados_pdf = tuple(tabla.encabezados[indice] for indice in indices_visibles)
    filas_pdf = tuple(
        tuple(fila[indice] for indice in indices_visibles)
        for fila in tabla.filas
    )

    def campo(etiqueta, valor):
        valor_visible = (
            "No registrado"
            if valor is None or str(valor).strip() == ""
            else valor
        )
        return Paragraph(
            f'<font size="7" color="#6B7280">{escape(str(etiqueta).upper())}</font>'
            f"<br/><b>{escape(str(valor_visible))}</b>",
            estilos["Campo"],
        )

    filas_metadata = []
    for indice in range(0, len(metadata), 3):
        fila = [campo(clave, valor) for clave, valor in metadata[indice : indice + 3]]
        while len(fila) < 3:
            fila.append("")
        filas_metadata.append(fila)
    tabla_metadata = Table(
        filas_metadata,
        colWidths=[ancho_util / 3] * 3,
    )
    tabla_metadata.setStyle(
        TableStyle(
            [
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, fondo_suave]),
                ("BOX", (0, 0), (-1, -1), 0.5, _color_tenue(color_acento, 0.65)),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, _color_tenue(color_acento, 0.75)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    filas_totales = []
    for indice in range(0, len(tabla.totales), 3):
        fila = [campo(clave, valor) for clave, valor in tabla.totales[indice : indice + 3]]
        while len(fila) < 3:
            fila.append("")
        filas_totales.append(fila)
    tabla_totales = Table(
        filas_totales or [[campo("Sin totales", "No aplica"), "", ""]],
        colWidths=[ancho_util / 3] * 3,
    )
    tabla_totales.setStyle(
        TableStyle(
            [
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, fondo_suave]),
                ("BOX", (0, 0), (-1, -1), 0.6, color_acento),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, _color_tenue(color_acento, 0.75)),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    encabezados_numericos = {
        "cantidad",
        "personas",
        "selecciones",
        "pagados",
        "pendientes",
        "otros",
        "pedidos",
        "ingresos",
        "subtotal",
        "descuentos",
        "impuestos",
        "impuesto",
        "envio",
        "total",
        "monto",
        "tasa",
        "base imponible",
    }
    indices_numericos = {
        indice
        for indice, encabezado in enumerate(encabezados_pdf)
        if str(encabezado).strip().lower() in encabezados_numericos
    }
    datos = [
        [Paragraph(escape(str(valor)), estilos["Encabezado"]) for valor in encabezados_pdf]
    ]
    for fila in filas_pdf:
        datos.append(
            [
                Paragraph(
                    escape(_limpiar_celda(valor)),
                    estilos["CeldaDerecha"] if indice in indices_numericos else estilos["Celda"],
                )
                for indice, valor in enumerate(fila)
            ]
        )

    pesos = []
    muestras = [encabezados_pdf, *filas_pdf[:50]]
    for indice in range(len(encabezados_pdf)):
        longitud = max(
            len(str(fila[indice] if indice < len(fila) else ""))
            for fila in muestras
        )
        pesos.append(max(8, min(longitud, 34)))
    total_pesos = sum(pesos) or 1
    anchos = [ancho_util * peso / total_pesos for peso in pesos]

    tabla_detalle = LongTable(
        datos,
        colWidths=anchos,
        repeatRows=1,
        splitByRow=1,
    )
    tabla_detalle.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), color_acento),
                ("BOX", (0, 0), (-1, -1), 0.6, color_acento),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, fondo_suave]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    elementos = [
        cabecera,
        Spacer(1, 4 * mm),
        banda,
        Spacer(1, 4 * mm),
        Paragraph("Datos del reporte", estilos["Seccion"]),
        tabla_metadata,
        Spacer(1, 3 * mm),
        KeepTogether(
            [
                Paragraph("Resumen de totales", estilos["Seccion"]),
                tabla_totales,
            ]
        ),
        Spacer(1, 4 * mm),
        Paragraph("Detalle", estilos["Seccion"]),
        tabla_detalle,
    ]

    def decorar_pagina(canvas, doc):
        margen = 14 * mm
        canvas.saveState()
        if canvas.getPageNumber() > 1:
            canvas.setStrokeColor(color_principal)
            canvas.setLineWidth(1)
            canvas.line(margen, alto_pagina - 12 * mm, ancho_pagina - margen, alto_pagina - 12 * mm)
            canvas.setFillColor(colors.HexColor("#4B5563"))
            canvas.setFont("Helvetica-Bold", 8)
            canvas.drawString(
                margen,
                alto_pagina - 9.5 * mm,
                f"{tabla.titulo} | {empresa.nombre}",
            )

        canvas.setStrokeColor(_color_tenue(color_principal, 0.7))
        canvas.setLineWidth(0.5)
        canvas.line(margen, 13 * mm, ancho_pagina - margen, 13 * mm)
        contacto = _texto_contacto_empresa(empresa) or empresa.nombre
        ancho_contacto = ancho_pagina - (58 * mm)
        tamano_contacto = _ajustar_fuente(contacto, ancho_contacto)
        canvas.setFillColor(colors.HexColor("#4B5563"))
        canvas.setFont("Helvetica", tamano_contacto)
        canvas.drawString(margen, 8.5 * mm, contacto)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.drawRightString(
            ancho_pagina - margen,
            8.5 * mm,
            f"Pagina {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    documento.build(
        elementos,
        onFirstPage=decorar_pagina,
        onLaterPages=decorar_pagina,
    )
    return salida.getvalue()
