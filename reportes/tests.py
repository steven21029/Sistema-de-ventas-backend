from datetime import UTC, datetime, timedelta
import re
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch
from zoneinfo import ZoneInfo

from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image as PILImage
from rest_framework import status
from rest_framework.test import APITestCase

from catalogo.models import Categoria, Familia, Producto
from empresas.models import Empresa, SucursalEmpresa
from pagos.models import Pago
from pedidos.models import DetallePedido, Pedido, Prefactura
from usuarios.models import PerfilUsuario


User = get_user_model()
ZONA_HONDURAS = ZoneInfo("America/Tegucigalpa")


class ReportesVentasAPITests(APITestCase):
    def setUp(self):
        self.analiza = Empresa.objects.create(
            nombre="Analiza",
            slug="analiza",
            modo_inventario=Empresa.ModoInventario.SIN_INVENTARIO,
            cobra_impuesto=True,
        )
        self.otra = Empresa.objects.create(
            nombre="Otra empresa",
            slug="otra",
            modo_inventario=Empresa.ModoInventario.SIN_INVENTARIO,
            cobra_impuesto=True,
        )
        self.admin = self._crear_usuario(
            "admin",
            PerfilUsuario.Rol.ADMINISTRADOR_EMPRESA,
            self.analiza,
        )
        self.gerente = self._crear_usuario(
            "gerente",
            PerfilUsuario.Rol.GERENTE,
            self.analiza,
        )
        self.comprador = self._crear_usuario(
            "comprador",
            PerfilUsuario.Rol.COMPRADOR,
            self.analiza,
        )
        self.comprador_otra = self._crear_usuario(
            "comprador-otra",
            PerfilUsuario.Rol.COMPRADOR,
            self.otra,
        )
        self.admin_otra = self._crear_usuario(
            "admin-otra",
            PerfilUsuario.Rol.ADMINISTRADOR_EMPRESA,
            self.otra,
        )
        self.maestro = self._crear_usuario(
            "maestro",
            PerfilUsuario.Rol.ADMINISTRADOR_MAESTRO,
            None,
        )
        self.maestro.perfil.empresas_permitidas.add(self.analiza)
        self.superusuario = User.objects.create_superuser(
            username="root",
            email="root@example.com",
            password="Prueba12345!",
        )

        familia = Familia.objects.create(empresa=self.analiza, nombre="Examenes")
        self.familia = familia
        categoria = Categoria.objects.create(
            empresa=self.analiza,
            familia=familia,
            nombre="Laboratorio",
        )
        self.hemograma = Producto.objects.create(
            empresa=self.analiza,
            familia=familia,
            categoria=categoria,
            codigo_barra="EXA-001",
            nombre="Hemograma",
            precio=Decimal("50.00"),
        )
        self.familia_imagenes = Familia.objects.create(
            empresa=self.analiza,
            nombre="Imagenes",
        )
        categoria_imagenes = Categoria.objects.create(
            empresa=self.analiza,
            familia=self.familia_imagenes,
            nombre="Ultrasonografia",
        )
        self.doppler = Producto.objects.create(
            empresa=self.analiza,
            familia=self.familia_imagenes,
            categoria=categoria_imagenes,
            codigo_barra="IMG-001",
            nombre="Doppler",
            precio=Decimal("50.00"),
        )
        self.sucursal_centro = SucursalEmpresa.objects.create(
            empresa=self.analiza,
            nombre="Sucursal Centro",
            ciudad="Tegucigalpa",
            direccion="Centro",
        )
        self.sucursal_norte = SucursalEmpresa.objects.create(
            empresa=self.analiza,
            nombre="Sucursal Norte",
            ciudad="San Pedro Sula",
            direccion="Norte",
        )

        self.pagado = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 5, 10, 0, tzinfo=ZONA_HONDURAS),
            estado=Pedido.EstadoPago.PAGADO,
            subtotal="100.00",
            descuentos="10.00",
            impuestos="13.50",
            envios="20.00",
            total="123.50",
        )
        DetallePedido.objects.create(
            pedido=self.pagado,
            producto=self.hemograma,
            precio_unitario=Decimal("50.00"),
            cantidad=2,
            porcentaje_descuento=10,
        )

        self.aprobado_por_pago = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 6, 11, 0, tzinfo=ZONA_HONDURAS),
            subtotal="50.00",
            impuestos="7.50",
            total="57.50",
        )
        pago_aprobado = Pago.objects.create(
            pedido=self.aprobado_por_pago,
            proveedor="prueba",
        )
        Pago.objects.filter(pk=pago_aprobado.pk).update(
            estado=Pago.Estado.APROBADO,
            fecha_confirmacion=self._utc(2026, 8, 6, 11, 5),
            fecha_creacion=self._utc(2026, 8, 6, 11, 1),
        )

        self.pendiente = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 7, 12, 0, tzinfo=ZONA_HONDURAS),
            subtotal="40.00",
            impuestos="6.00",
            total="46.00",
        )
        self.rechazado = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 8, 13, 0, tzinfo=ZONA_HONDURAS),
            subtotal="30.00",
            impuestos="4.50",
            total="34.50",
        )
        pago_rechazado = Pago.objects.create(
            pedido=self.rechazado,
            proveedor="prueba",
        )
        Pago.objects.filter(pk=pago_rechazado.pk).update(
            estado=Pago.Estado.RECHAZADO,
            fecha_confirmacion=self._utc(2026, 8, 8, 13, 5),
            fecha_creacion=self._utc(2026, 8, 8, 13, 1),
        )
        self.cancelado = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 9, 14, 0, tzinfo=ZONA_HONDURAS),
            estado=Pedido.EstadoPago.CANCELADO,
            subtotal="20.00",
            impuestos="3.00",
            total="23.00",
        )
        self.anterior = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 7, 10, 10, 0, tzinfo=ZONA_HONDURAS),
            estado=Pedido.EstadoPago.PAGADO,
            subtotal="80.00",
            impuestos="12.00",
            total="92.00",
        )
        self.fuera_rango = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 9, 1, 0, 1, tzinfo=ZONA_HONDURAS),
            estado=Pedido.EstadoPago.PAGADO,
            subtotal="200.00",
            impuestos="30.00",
            total="230.00",
        )
        self.pedido_otra = self._crear_pedido(
            self.otra,
            self.comprador_otra,
            datetime(2026, 8, 5, 10, 0, tzinfo=ZONA_HONDURAS),
            estado=Pedido.EstadoPago.PAGADO,
            subtotal="999.00",
            impuestos="149.85",
            total="1148.85",
        )

    def _crear_usuario(self, username, rol, empresa):
        usuario = User.objects.create_user(
            username=username,
            email=f"{username}@example.com",
            password="Prueba12345!",
        )
        perfil = usuario.perfil
        perfil.rol = rol
        perfil.empresa = empresa
        perfil.correo_verificado = True
        perfil.activo = True
        perfil.save()
        return usuario

    def _crear_pedido(
        self,
        empresa,
        usuario,
        fecha,
        estado=Pedido.EstadoPago.PENDIENTE,
        subtotal="100.00",
        descuentos="0.00",
        impuestos="15.00",
        envios="0.00",
        total="115.00",
    ):
        pedido = Pedido.objects.create(
            empresa=empresa,
            usuario=usuario,
            tipo_entrega=Pedido.TipoEntrega.RETIRO_EN_LOCAL,
            subtotal=Decimal(subtotal),
            descuento_total=Decimal(descuentos),
        )
        Pedido.objects.filter(pk=pedido.pk).update(
            estado_pago=estado,
            subtotal=Decimal(subtotal),
            descuento_total=Decimal(descuentos),
            impuesto=Decimal(impuestos),
            envio=Decimal(envios),
            total=Decimal(total),
            fecha_creacion=fecha.astimezone(UTC),
        )
        return Pedido.objects.get(pk=pedido.pk)

    def _utc(self, year, month, day, hour, minute):
        return datetime(
            year,
            month,
            day,
            hour,
            minute,
            tzinfo=ZONA_HONDURAS,
        ).astimezone(UTC)

    def _parametros_resumen(self, **extra):
        parametros = {
            "empresa_slug": self.analiza.slug,
            "fecha_desde": "2026-08-01",
            "fecha_hasta": "2026-08-31",
            "agrupacion": "mes",
            "comparar_periodo_anterior": "true",
        }
        parametros.update(extra)
        return parametros

    def _parametros_exportacion(self, formato, tipo):
        return {
            "empresa_slug": self.analiza.slug,
            "fecha_desde": "2026-08-01",
            "fecha_hasta": "2026-08-31",
            "formato": formato,
            "tipo": tipo,
        }

    def test_resumen_usa_solo_ventas_confirmadas_y_totales_historicos(self):
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        resumen = respuesta.data["resumen"]
        self.assertNotIn("empresa_slug", respuesta.data)
        self.assertNotIn("moneda", respuesta.data)
        self.assertEqual(resumen["ingresos_confirmados"], "181.00")
        self.assertEqual(resumen["ventas_confirmadas"], 2)
        self.assertEqual(resumen["ticket_promedio"], "90.50")
        self.assertEqual(resumen["subtotal"], "150.00")
        self.assertEqual(resumen["descuentos"], "10.00")
        self.assertEqual(resumen["impuestos"], "21.00")
        self.assertEqual(resumen["envios"], "20.00")
        self.assertEqual(resumen["monto_pendiente"], "46.00")
        self.assertEqual(resumen["pedidos_pendientes"], 1)
        self.assertEqual(
            resumen["pendientes_por_metodo"],
            {
                "sucursal": {"cantidad": 0, "monto": "0.00"},
                "en_linea": {"cantidad": 0, "monto": "0.00"},
                "sin_metodo": {"cantidad": 2, "monto": "80.50"},
            },
        )

    def test_pagos_por_metodo_cuenta_confirmados_y_deduplica_pedidos(self):
        pago_sucursal = Pago.objects.create(
            pedido=self.pagado,
            proveedor="sucursal",
            metodo=Pago.Metodo.SUCURSAL,
        )
        Pago.objects.filter(pk=pago_sucursal.pk).update(
            estado=Pago.Estado.APROBADO,
            fecha_confirmacion=self._utc(2026, 8, 5, 10, 5),
        )
        pago_duplicado = Pago.objects.create(
            pedido=self.aprobado_por_pago,
            proveedor="reintento",
            metodo=Pago.Metodo.EN_LINEA,
        )
        Pago.objects.filter(pk=pago_duplicado.pk).update(
            estado=Pago.Estado.APROBADO,
            fecha_confirmacion=self._utc(2026, 8, 6, 11, 10),
        )
        Pago.objects.create(
            pedido=self.pendiente,
            proveedor="sucursal",
            metodo=Pago.Metodo.SUCURSAL,
        )

        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta.data["resumen"]["pagos_por_metodo"],
            {
                "sucursal": {"cantidad": 1, "monto": "123.50"},
                "en_linea": {"cantidad": 1, "monto": "57.50"},
            },
        )

    def test_pagos_por_metodo_devuelve_ambos_en_cero_sin_pagos_confirmados(self):
        Pedido.objects.filter(pk=self.pagado.pk).update(
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
        )
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(
                fecha_desde="2026-08-05",
                fecha_hasta="2026-08-05",
                comparar_periodo_anterior="false",
            ),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta.data["resumen"]["pagos_por_metodo"],
            {
                "sucursal": {"cantidad": 0, "monto": "0.00"},
                "en_linea": {"cantidad": 0, "monto": "0.00"},
            },
        )

    def test_pagos_por_metodo_respeta_empresa_y_rango_del_pedido(self):
        for pedido, metodo in (
            (self.pedido_otra, Pago.Metodo.SUCURSAL),
            (self.fuera_rango, Pago.Metodo.EN_LINEA),
        ):
            pago = Pago.objects.create(
                pedido=pedido,
                proveedor="fuera-reporte",
                metodo=metodo,
            )
            Pago.objects.filter(pk=pago.pk).update(
                estado=Pago.Estado.APROBADO,
                fecha_confirmacion=self._utc(2026, 8, 10, 10, 0),
            )

        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta.data["resumen"]["pagos_por_metodo"],
            {
                "sucursal": {"cantidad": 0, "monto": "0.00"},
                "en_linea": {"cantidad": 1, "monto": "57.50"},
            },
        )

    def test_pendientes_por_metodo_agrupa_pedidos_sin_duplicar_intentos(self):
        pendiente_sucursal = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 11, 10, 0, tzinfo=ZONA_HONDURAS),
            total="70.00",
        )
        pendiente_linea = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 12, 10, 0, tzinfo=ZONA_HONDURAS),
            total="80.00",
        )
        Pedido.objects.filter(pk=pendiente_sucursal.pk).update(
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
        )
        Pedido.objects.filter(pk=pendiente_linea.pk).update(
            metodo_pago=Pedido.MetodoPago.EN_LINEA,
        )
        Pago.objects.create(
            pedido=pendiente_sucursal,
            proveedor="sucursal",
            metodo=Pago.Metodo.SUCURSAL,
        )
        intento_rechazado = Pago.objects.create(
            pedido=pendiente_linea,
            proveedor="primer-intento",
            metodo=Pago.Metodo.EN_LINEA,
        )
        Pago.objects.filter(pk=intento_rechazado.pk).update(
            estado=Pago.Estado.RECHAZADO,
        )
        Pago.objects.create(
            pedido=pendiente_linea,
            proveedor="segundo-intento",
            metodo=Pago.Metodo.EN_LINEA,
        )

        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta.data["resumen"]["pendientes_por_metodo"],
            {
                "sucursal": {"cantidad": 1, "monto": "70.00"},
                "en_linea": {"cantidad": 1, "monto": "80.00"},
                "sin_metodo": {"cantidad": 2, "monto": "80.50"},
            },
        )

    def test_pendientes_por_metodo_respeta_empresa_y_rango(self):
        pendiente_otra = self._crear_pedido(
            self.otra,
            self.comprador_otra,
            datetime(2026, 8, 15, 10, 0, tzinfo=ZONA_HONDURAS),
            total="500.00",
        )
        pendiente_fuera = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 9, 2, 10, 0, tzinfo=ZONA_HONDURAS),
            total="600.00",
        )
        Pedido.objects.filter(pk=pendiente_otra.pk).update(
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
        )
        Pedido.objects.filter(pk=pendiente_fuera.pk).update(
            metodo_pago=Pedido.MetodoPago.EN_LINEA,
        )

        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta.data["resumen"]["pendientes_por_metodo"],
            {
                "sucursal": {"cantidad": 0, "monto": "0.00"},
                "en_linea": {"cantidad": 0, "monto": "0.00"},
                "sin_metodo": {"cantidad": 2, "monto": "80.50"},
            },
        )

    def test_reporte_vence_prefactura_y_la_excluye_de_pendientes(self):
        pendiente_sucursal = self._crear_pedido(
            self.analiza,
            self.comprador,
            datetime(2026, 8, 11, 10, 0, tzinfo=ZONA_HONDURAS),
            total="70.00",
        )
        pendiente_sucursal.seleccionar_metodo_pago(
            Pedido.MetodoPago.SUCURSAL,
            sucursal=self.sucursal_centro,
        )
        pago = Pago.objects.create(
            pedido=pendiente_sucursal,
            proveedor="sucursal",
            metodo=Pago.Metodo.SUCURSAL,
        )
        prefactura = Prefactura.obtener_o_crear_para_pedido(pendiente_sucursal)
        Prefactura.objects.filter(pk=prefactura.pk).update(
            fecha_vencimiento=timezone.now() - timedelta(seconds=1)
        )

        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta.data["resumen"]["pendientes_por_metodo"]["sucursal"],
            {"cantidad": 0, "monto": "0.00"},
        )
        estados = {item["estado"]: item for item in respuesta.data["estados"]}
        self.assertEqual(estados["rechazado"]["cantidad"], 2)
        pendiente_sucursal.refresh_from_db()
        pago.refresh_from_db()
        self.assertEqual(
            pendiente_sucursal.estado_pago,
            Pedido.EstadoPago.RECHAZADO,
        )
        self.assertEqual(pago.estado, Pago.Estado.RECHAZADO)

    def test_desglosa_pagados_pendientes_rechazados_y_cancelados(self):
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        estados = {item["estado"]: item for item in respuesta.data["estados"]}
        self.assertEqual(estados["pagado"], {"estado": "pagado", "cantidad": 2, "monto": "181.00"})
        self.assertEqual(estados["pendiente"]["cantidad"], 1)
        self.assertEqual(estados["rechazado"]["cantidad"], 1)
        self.assertEqual(estados["cancelado"]["cantidad"], 1)

    def test_comparacion_mensual_y_producto_mas_vendido(self):
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )

        self.assertEqual(
            respuesta.data["resumen"]["variacion_ingresos_porcentaje"],
            96.7,
        )
        self.assertEqual(
            respuesta.data["resumen"]["variacion_ventas_porcentaje"],
            100.0,
        )
        self.assertEqual(
            respuesta.data["serie"],
            [{"periodo": "2026-08", "etiqueta": "Ago", "ingresos": "181.00", "ventas": 2}],
        )
        self.assertEqual(
            respuesta.data["productos_mas_vendidos"][0],
            {
                "codigo": "EXA-001",
                "nombre": "Hemograma",
                "cantidad": 2,
                "ingresos": "90.00",
            },
        )

    def test_rango_fechas_es_inclusivo_y_respeta_hora_honduras(self):
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(
                fecha_desde="2026-08-05",
                fecha_hasta="2026-08-05",
                agrupacion="dia",
                comparar_periodo_anterior="false",
            ),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(respuesta.data["resumen"]["ventas_confirmadas"], 1)
        self.assertEqual(respuesta.data["resumen"]["ingresos_confirmados"], "123.50")
        self.assertEqual(respuesta.data["serie"][0]["periodo"], "2026-08-05")

    def test_rechaza_rango_invertido(self):
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(
                fecha_desde="2026-08-31",
                fecha_hasta="2026-08-01",
            ),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("fecha_hasta", respuesta.data)

    def test_filtros_por_ciudad_sucursal_examen_y_familia(self):
        Pedido.objects.filter(pk=self.pagado.pk).update(
            municipio_entrega="Tegucigalpa",
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
            sucursal_pago=self.sucursal_centro,
        )
        Pedido.objects.filter(pk=self.aprobado_por_pago.pk).update(
            municipio_entrega="San Pedro Sula",
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
            sucursal_pago=self.sucursal_norte,
        )
        DetallePedido.objects.create(
            pedido=self.aprobado_por_pago,
            producto=self.doppler,
            precio_unitario=Decimal("50.00"),
            cantidad=1,
        )
        self.client.force_authenticate(self.admin)

        casos = (
            ({"ciudad": "Tegucigalpa"}, "123.50"),
            ({"sucursal_id": self.sucursal_centro.pk}, "123.50"),
            ({"examen_id": self.hemograma.pk}, "123.50"),
            ({"familia_id": self.familia_imagenes.pk}, "57.50"),
        )
        for filtros, ingreso_esperado in casos:
            with self.subTest(filtros=filtros):
                respuesta = self.client.get(
                    reverse("reportes-resumen-ventas"),
                    self._parametros_resumen(**filtros),
                )
                self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
                self.assertEqual(
                    respuesta.data["resumen"]["ingresos_confirmados"],
                    ingreso_esperado,
                )
                for clave, valor in filtros.items():
                    self.assertEqual(
                        respuesta.data["filtros_aplicados"][clave],
                        valor,
                    )

        por_examen = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(examen_id=self.hemograma.pk),
        )
        self.assertEqual(
            [item["nombre"] for item in por_examen.data["productos_mas_vendidos"]],
            ["Hemograma"],
        )

    def test_filtros_rechazan_ids_ajenos_o_incompatibles(self):
        familia_otra = Familia.objects.create(
            empresa=self.otra,
            nombre="Familia ajena",
        )
        categoria_otra = Categoria.objects.create(
            empresa=self.otra,
            familia=familia_otra,
            nombre="Categoria ajena",
        )
        producto_otra = Producto.objects.create(
            empresa=self.otra,
            familia=familia_otra,
            categoria=categoria_otra,
            codigo_barra="OTRA-001",
            nombre="Producto ajeno",
            precio=Decimal("10.00"),
        )
        sucursal_otra = SucursalEmpresa.objects.create(
            empresa=self.otra,
            nombre="Sucursal ajena",
            ciudad="La Ceiba",
            direccion="Otra direccion",
        )
        self.client.force_authenticate(self.admin)

        for filtros, campo in (
            ({"sucursal_id": sucursal_otra.pk}, "sucursal_id"),
            ({"examen_id": producto_otra.pk}, "examen_id"),
            ({"familia_id": familia_otra.pk}, "familia_id"),
            (
                {
                    "examen_id": self.hemograma.pk,
                    "familia_id": self.familia_imagenes.pk,
                },
                "examen_id",
            ),
        ):
            with self.subTest(filtros=filtros):
                respuesta = self.client.get(
                    reverse("reportes-resumen-ventas"),
                    self._parametros_resumen(**filtros),
                )
                self.assertEqual(
                    respuesta.status_code,
                    status.HTTP_400_BAD_REQUEST,
                )
                self.assertIn(campo, respuesta.data)

    def test_reporte_sucursales_ordena_personas_y_selecciones(self):
        segundo_comprador = self._crear_usuario(
            "segundo-comprador",
            PerfilUsuario.Rol.COMPRADOR,
            self.analiza,
        )
        segundo_pedido = self._crear_pedido(
            self.analiza,
            segundo_comprador,
            datetime(2026, 8, 10, 9, 0, tzinfo=ZONA_HONDURAS),
            estado=Pedido.EstadoPago.PAGADO,
            total="70.00",
        )
        Pedido.objects.filter(pk__in=[self.pagado.pk, self.pendiente.pk]).update(
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
            sucursal_pago=self.sucursal_centro,
        )
        Pedido.objects.filter(pk=segundo_pedido.pk).update(
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
            sucursal_pago=self.sucursal_centro,
        )
        Pedido.objects.filter(pk=self.rechazado.pk).update(
            metodo_pago=Pedido.MetodoPago.SUCURSAL,
            sucursal_pago=self.sucursal_norte,
        )
        self.client.force_authenticate(self.admin)

        respuesta = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("xlsx", "sucursales"),
        )
        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        hoja = load_workbook(BytesIO(respuesta.content), data_only=True).active
        filas = list(hoja.iter_rows(values_only=True))
        fila_encabezados = next(
            indice for indice, fila in enumerate(filas) if "Personas" in fila
        )
        detalle = [
            fila[:8]
            for fila in filas[fila_encabezados + 1 :]
            if fila[1] in {"Sucursal Centro", "Sucursal Norte"}
        ]
        self.assertEqual(detalle[0][1], "Sucursal Centro")
        self.assertEqual(detalle[0][2:4], (2, 3))
        self.assertEqual(detalle[1][1], "Sucursal Norte")
        self.assertEqual(detalle[1][2:4], (1, 1))

        pdf = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("pdf", "sucursales"),
        )
        self.assertEqual(pdf.status_code, status.HTTP_200_OK)
        self.assertIn(b"Visitas estimadas por sucursal", pdf.content)

    def test_reporte_familias_usa_solo_ventas_confirmadas(self):
        producto_sin_ventas = Producto.objects.create(
            empresa=self.analiza,
            familia=self.familia_imagenes,
            categoria=self.doppler.categoria,
            codigo_barra="IMG-002",
            nombre="Ultrasonografia sin ventas",
            precio=Decimal("75.00"),
        )
        DetallePedido.objects.create(
            pedido=self.aprobado_por_pago,
            producto=self.doppler,
            precio_unitario=Decimal("50.00"),
            cantidad=1,
        )
        DetallePedido.objects.create(
            pedido=self.pendiente,
            producto=self.doppler,
            precio_unitario=Decimal("50.00"),
            cantidad=1,
        )
        self.client.force_authenticate(self.admin)
        parametros = self._parametros_exportacion("xlsx", "familias") | {
            "familia_id": self.familia_imagenes.pk,
        }

        respuesta = self.client.get(
            reverse("reportes-ventas-exportar"),
            parametros,
        )
        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        hoja = load_workbook(BytesIO(respuesta.content), data_only=True).active
        filas = list(hoja.iter_rows(values_only=True))
        valores = [celda for fila in filas for celda in fila if celda is not None]
        self.assertIn("Detalle de productos por familia", valores)
        self.assertIn("Imagenes", valores)
        self.assertIn("Familia", valores)
        self.assertIn("Doppler", valores)
        self.assertIn(producto_sin_ventas.nombre, valores)
        self.assertNotIn("Examenes", valores)
        fila_encabezados = next(
            indice
            for indice, fila in enumerate(filas)
            if "Producto o examen" in fila
        )
        detalle = {
            fila[2]: fila
            for fila in filas[fila_encabezados + 1 :]
            if fila[2] in {"Doppler", producto_sin_ventas.nombre}
        }
        self.assertEqual(detalle["Doppler"][4:], (1, 1, 50))
        self.assertEqual(
            detalle[producto_sin_ventas.nombre][4:],
            (0, 0, 0),
        )

        for ruta in (
            "/api/reportes/ventas/exportar/",
            "/api/v1/reportes/ventas/exportar/",
        ):
            pdf = self.client.get(
                ruta,
                self._parametros_exportacion("pdf", "familias"),
            )
            self.assertEqual(pdf.status_code, status.HTTP_200_OK)
            self.assertIn(b"Detalle de productos por familia", pdf.content)

    def test_aislamiento_multiempresa(self):
        self.client.force_authenticate(self.admin)
        propia = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(),
        )
        ajena = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(empresa_slug=self.otra.slug),
        )

        self.assertEqual(propia.status_code, status.HTTP_200_OK)
        self.assertNotEqual(
            propia.data["resumen"]["ingresos_confirmados"],
            "1148.85",
        )
        self.assertEqual(ajena.status_code, status.HTTP_403_FORBIDDEN)

    def test_permisos_por_rol(self):
        url = reverse("reportes-resumen-ventas")
        for usuario in (self.superusuario, self.admin, self.gerente, self.maestro):
            self.client.force_authenticate(usuario)
            respuesta = self.client.get(url, self._parametros_resumen())
            self.assertEqual(respuesta.status_code, status.HTTP_200_OK)

        for usuario in (self.comprador, self.admin_otra):
            self.client.force_authenticate(usuario)
            respuesta = self.client.get(url, self._parametros_resumen())
            self.assertEqual(respuesta.status_code, status.HTTP_403_FORBIDDEN)

        self.client.force_authenticate(self.comprador)
        exportacion = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("xlsx", "resumen"),
        )
        self.assertEqual(exportacion.status_code, status.HTTP_403_FORBIDDEN)

    def test_maestro_no_accede_empresa_no_asignada(self):
        self.client.force_authenticate(self.maestro)
        respuesta = self.client.get(
            reverse("reportes-resumen-ventas"),
            self._parametros_resumen(empresa_slug=self.otra.slug),
        )
        self.assertEqual(respuesta.status_code, status.HTTP_403_FORBIDDEN)

    def test_rutas_api_y_api_v1_son_compatibles(self):
        self.client.force_authenticate(self.admin)
        for ruta in (
            "/api/reportes/resumen-ventas/",
            "/api/v1/reportes/resumen-ventas/",
        ):
            respuesta = self.client.get(ruta, self._parametros_resumen())
            self.assertEqual(respuesta.status_code, status.HTTP_200_OK)

        for ruta in (
            "/api/reportes/ventas/exportar/",
            "/api/v1/reportes/ventas/exportar/",
        ):
            respuesta = self.client.get(
                ruta,
                self._parametros_exportacion("xlsx", "resumen"),
            )
            self.assertEqual(respuesta.status_code, status.HTTP_200_OK)

    def test_exportacion_csv_ya_no_esta_disponible(self):
        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("csv", "ventas"),
        )

        self.assertEqual(respuesta.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("formato", respuesta.data)

    def test_exportacion_xlsx(self):
        self.analiza.telefono = "22334455"
        self.analiza.correo = "reportes@analiza.test"
        self.analiza.direccion = "Boulevard principal"
        self.analiza.sitio_web = "https://analiza.test"
        self.analiza.save()
        logo = BytesIO()
        PILImage.new("RGB", (320, 120), "white").save(logo, format="PNG")
        self.client.force_authenticate(self.admin)
        with patch(
            "reportes.services._contenido_logo_empresa",
            return_value=logo.getvalue(),
        ):
            respuesta = self.client.get(
                reverse("reportes-ventas-exportar"),
                self._parametros_exportacion("xlsx", "pagos"),
            )

        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertEqual(
            respuesta["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        libro = load_workbook(BytesIO(respuesta.content))
        hoja = libro.active
        valores = [
            str(celda)
            for fila in hoja.iter_rows(values_only=True)
            for celda in fila
            if celda is not None
        ]
        self.assertIn("Analiza", valores)
        self.assertIn("Detalle de pagos", valores)
        self.assertNotIn("Empresa slug", valores)
        self.assertNotIn("Moneda", valores)
        self.assertNotIn("HNL", valores)
        self.assertIn("REPORTE COMERCIAL", valores)
        self.assertIn("RESUMEN DE TOTALES", valores)
        self.assertIn("DETALLE", valores)
        self.assertTrue(any("22334455" in valor for valor in valores))
        self.assertFalse(hoja.sheet_view.showGridLines)
        self.assertEqual(hoja.page_setup.orientation, "landscape")
        self.assertEqual(hoja.page_setup.paperSize, 9)
        self.assertEqual(hoja.page_setup.fitToWidth, 1)
        self.assertIsNotNone(hoja.freeze_panes)
        self.assertTrue(hoja.auto_filter.ref)
        self.assertTrue(hoja.merged_cells.ranges)
        self.assertEqual(len(hoja._images), 1)

        fila_encabezados = next(
            celda.row
            for fila in hoja.iter_rows()
            for celda in fila
            if celda.value == "Monto"
        )
        columna_monto = next(
            celda.column
            for celda in hoja[fila_encabezados]
            if celda.value == "Monto"
        )
        montos = [
            hoja.cell(row=fila, column=columna_monto).value
            for fila in range(fila_encabezados + 1, hoja.max_row + 1)
            if hoja.cell(row=fila, column=columna_monto).value is not None
        ]
        self.assertTrue(any(isinstance(valor, (int, float)) for valor in montos))
        self.assertTrue(
            hoja.cell(row=fila_encabezados, column=1).fill.fgColor.rgb.endswith(
                "2D4B77"
            )
        )
        self.assertEqual(
            respuesta["Content-Disposition"],
            'attachment; filename="reporte_pagos_2026-08-01_2026-08-31.xlsx"',
        )

        resumen = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("xlsx", "resumen"),
        )
        valores_resumen = [
            celda
            for fila in load_workbook(BytesIO(resumen.content), read_only=True)
            .active.iter_rows(values_only=True)
            for celda in fila
            if celda is not None
        ]
        self.assertNotIn("Estado", valores_resumen)
        self.assertIn("Producto", valores_resumen)

    def test_exportacion_pdf(self):
        self.analiza.telefono = "22334455"
        self.analiza.correo = "reportes@analiza.test"
        self.analiza.direccion = "Boulevard principal"
        self.analiza.sitio_web = "https://analiza.test"
        self.analiza.save()
        self.client.force_authenticate(self.admin)
        titulos = {
            "resumen": b"Resumen comercial",
            "ventas": b"Detalle de ventas",
            "pagos": b"Detalle de pagos",
            "impuestos": b"Detalle de impuestos",
        }

        for tipo, titulo in titulos.items():
            with self.subTest(tipo=tipo):
                respuesta = self.client.get(
                    reverse("reportes-ventas-exportar"),
                    self._parametros_exportacion("pdf", tipo),
                )

                self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
                self.assertEqual(respuesta["Content-Type"], "application/pdf")
                self.assertIn(".pdf", respuesta["Content-Disposition"])
                self.assertEqual(
                    respuesta["Content-Disposition"],
                    (
                        f'attachment; filename="reporte_{tipo}_'
                        '2026-08-01_2026-08-31.pdf"'
                    ),
                )
                self.assertTrue(respuesta.content.startswith(b"%PDF"))
                self.assertRegex(
                    respuesta.content,
                    rb"/MediaBox\s*\[\s*0\s+0\s+841\.8898\s+595\.2756\s*\]",
                )
                for texto in (
                    b"REPORTE COMERCIAL",
                    titulo,
                    b"Analiza",
                    b"22334455",
                    b"reportes@analiza.test",
                    b"Boulevard principal",
                    b"https://analiza.test",
                    b"Resumen de totales",
                    b"Detalle",
                ):
                    self.assertIn(texto, respuesta.content)
                self.assertNotIn(b"EMPRESA SLUG", respuesta.content)
                self.assertNotIn(b"MONEDA", respuesta.content)
                self.assertNotIn(b"HNL", respuesta.content)

        respuesta_vacia = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("pdf", "resumen")
            | {"fecha_desde": "2025-01-01", "fecha_hasta": "2025-01-31"},
        )
        self.assertEqual(respuesta_vacia.status_code, status.HTTP_200_OK)
        self.assertNotIn(b"No registrado", respuesta_vacia.content)

    def test_exportacion_pdf_de_varias_paginas_conserva_detalle_y_paginacion(self):
        ultimo = None
        for _indice in range(65):
            ultimo = self._crear_pedido(
                self.analiza,
                self.comprador,
                datetime(2026, 8, 15, 9, 0, tzinfo=ZONA_HONDURAS),
            )

        self.client.force_authenticate(self.admin)
        respuesta = self.client.get(
            reverse("reportes-ventas-exportar"),
            self._parametros_exportacion("pdf", "ventas"),
        )

        paginas = re.findall(rb"/Type\s*/Page(?!s)", respuesta.content)
        self.assertEqual(respuesta.status_code, status.HTTP_200_OK)
        self.assertGreater(len(paginas), 1)
        self.assertIn(ultimo.numero.encode(), respuesta.content)
        self.assertIn(b"Pagina 2", respuesta.content)
